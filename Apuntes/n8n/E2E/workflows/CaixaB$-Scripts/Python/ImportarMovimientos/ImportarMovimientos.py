
import pandas as pd
from openpyxl import load_workbook
from copy import copy
from pathlib import Path

# ====================================================================
# 1. CONFIGURACIÓN
#    (equivalente al "formulario de variables" del script original)
# ====================================================================

CONFIG = { #Diccionario con las variables globales
    # Archivo histórico donde se acumulan todos los movimientos ya importados
    # (equivalente a idSpreadSheet + gidHojaHistorico)
    "ruta_historico": Path("C:/Users/Oscar Ardevol/Documents/MovimientosBancarios/Norgenic/€/Historico_py.xlsx"),
    "hoja_historico": "BD_Banco",
    "hoja_movimientos_cuenta": "Movimientos_cuenta_0087231",

    # Carpeta donde dejamos el Excel bruto exportado por el banco
    # (equivalente a idCarpetaDrive)
    "carpeta_importar": Path("C:/Users/Oscar Ardevol/Documents/MovimientosBancarios/Norgenic/€/Por archivar"),
    "carpeta_archivados": Path("C:/Users/Oscar Ardevol/Documents/MovimientosBancarios/Norgenic/€/Archivados"),

    # Fila (1-indexada, como en Sheets/Excel) donde empiezan los datos útiles.
    # Varía entre bancos -> equivalente a filaInicioDatosImportados
    "fila_inicio_datos": 4,

    # Nº de columnas de origen que nos interesan: Fecha, FechaValor,
    # Movimiento, MasDatos, Importe, Saldo (A:F)
    "num_columnas_origen": 6,
}


# ====================================================================
# 2. GENERACIÓN DE UID
#    Réplica exacta de la lógica del bloque `const uids = inverseBD.map(...)`
# ====================================================================

def formatear_fecha(valor) -> str:
    """Convierte una fecha a texto DD/MM/YYYY.
    Si está vacía devuelve "" (igual que el `row[0] ? ... : ''` original)."""
    if pd.isna(valor) or valor == "":
        return ""
    if isinstance(valor, str):
        try:
            valor = pd.to_datetime(valor, dayfirst=True)
        except Exception:
            # Si no es una fecha reconocible, lo dejamos tal cual (best-effort,
            # igual que haría Apps Script con new Date() sobre texto raro).
            return valor
    return f"{valor.day:02d}/{valor.month:02d}/{valor.year}"


def formatear_importe(valor) -> str:
    r"""Réplica de: row.toString().replace(/,/g,'').replace(/\./g,',').
    En JavaScript un número entero 1414.0 se convierte en "1414".
    """
    if pd.isna(valor) or valor == "":
        return ""

    if isinstance(valor, (int, float)):
        if float(valor).is_integer():
            return str(int(valor))
        return str(valor).replace(".", ",")

    texto = str(valor).strip()
    if texto == "":
        return ""

    texto_sin_miles = texto.replace(",", "")
    try:
        valor_float = float(texto_sin_miles)
        if valor_float.is_integer():
            return str(int(valor_float))
    except ValueError:
        pass

    return texto.replace(".", ",")


def generar_uid(fila: list) -> str:
    """Genera el UID concatenando los 6 campos con el separador "'_'"."""
    fecha = formatear_fecha(fila[0])
    fecha_valor = formatear_fecha(fila[1])
    movimiento = str(fila[2]).strip() if not pd.isna(fila[2]) else ""
    mas_datos = str(fila[3]).strip() if not pd.isna(fila[3]) else ""
    importe = formatear_importe(fila[4])
    saldo = formatear_importe(fila[5])
    return "'_'".join([fecha, fecha_valor, movimiento, mas_datos, importe, saldo])


# ====================================================================
# 3. LECTURA DEL ARCHIVO BANCARIO BRUTO
#    Equivalente a getMovimientosBancarios()
# ====================================================================

def obtener_primer_excel(carpeta: Path) -> Path:
    """Busca el primer .xlsx o .xls en la carpeta de importación.
    Equivalente a folderMovimientosBancarios.getFilesByType(...).next()"""
    archivos = sorted(
        [p for p in carpeta.iterdir() if p.suffix.lower() in {".xlsx", ".xls"}]
    )
    if not archivos:
        raise FileNotFoundError(f"No se encontró ningún .xlsx/.xls en {carpeta}")
    return archivos[0]


def leer_movimientos_bancarios(config: dict) -> list:
    """Lee el archivo bruto, invierte el orden, filtra filas vacías y genera
    el UID de cada movimiento.
    Devuelve una lista de filas: [UID, Fecha, FechaValor, Movimiento, MasDatos, Importe, Saldo]
    """
    archivo = obtener_primer_excel(config["carpeta_importar"])
    print(f"Leyendo archivo: {archivo.name}")

    # header=None: no confiamos en que la fila de cabecera del banco sea fiable,
    # igual que el original que empieza a leer directamente en filaInicioDatosImportados.
    engine = "openpyxl" if archivo.suffix.lower() == ".xlsx" else "xlrd"
    df = pd.read_excel(
        archivo,
        header=None,
        skiprows=config["fila_inicio_datos"] - 1,
        engine=engine, #Le dice a pandas como (que libreria se usa) y bajo que condiciones abrir el excel
    )

    # Nos quedamos solo con las columnas que nos interesan (A:F),
    # por si el banco añade columnas extra a la derecha.
    df = df.iloc[:, :config["num_columnas_origen"]] #Sacamos con iloc rango de filas, columnas ej: [0:4 , 0:6]

    # Filtramos filas totalmente vacías / sin fecha en la columna A
    # (equivalente a: inverseBD.filter(row => row[0] !== ""))
    df = df[df.iloc[:, 0].notna()]

    filas = df.values.tolist()

    # Invertimos el orden (el banco exporta normalmente del más reciente
    # al más antiguo; queremos guardar en orden cronológico ascendente)
    filas.reverse()

    # Usamos un bulce para recorer cada fila, generar el UID y lo anteponemos a cada fila
    filas_con_uid = [[generar_uid(fila)] + fila for fila in filas]

    return filas_con_uid


# ====================================================================
# 4. COMPARACIÓN CON EL HISTÓRICO Y CARGA DE MOVIMIENTOS NUEVOS
#    Equivalente a appendBD()
# ====================================================================

def leer_uids_existentes(config: dict) -> set:
    """Lee la columna UID (columna A) del histórico para poder detectar
    duplicados, igual que bdSheetRangeUIDs + historicoUIDs en el original."""
    wb = load_workbook(config["ruta_historico"], read_only=True, data_only=True)
    ws = wb[config["hoja_historico"]]

    uids = set() # Hace una lista de los UID del historico (BD_Banco) para tratarlas en memoria sin tocar el archivo
    for fila in ws.iter_rows(min_row=2, max_col=1, values_only=True):  # min_row=2 salta cabecera
        if fila[0] is not None:
            uids.add(str(fila[0]).strip())

    wb.close()
    return uids

# Esta función se  podra sustituir por "lastRow_BDB= len(uids)" cuando no haya posibilidad de UID's duplicadas.
def encontrar_siguiente_fila(ws) -> int:
    """Devuelve la fila siguiente a la última fila con contenido."""
    for row_idx in range(ws.max_row, 0, -1):
        hay_datos = any(
            ws.cell(row=row_idx, column=col_idx).value is not None
            and str(ws.cell(row=row_idx, column=col_idx).value).strip() != ""
            for col_idx in range(1, ws.max_column + 1)
        )
        if hay_datos:
            return row_idx + 1
    return 1

#Ese valores viene de la llamada: escribir_fila_con_formato(sheet_BDB, fila) y fila es cada elemento de filas_a_importar.
def escribir_fila_con_formato(ws, valores: list) -> None: 
    """Escribe una fila en la siguiente línea libre y copia el formato de la fila anterior."""
    fila_destino = encontrar_siguiente_fila(ws)
    fila_origen = fila_destino - 1 if fila_destino > 1 else 1

    for col_idx, valor in enumerate(valores, start=1): # Itera cada columna y fila para pegarle los valores importados
        celda_destino = ws.cell(row=fila_destino, column=col_idx) # LastRow=i; columna = i
        celda_destino.value = valor # La linia anterior recorre cada fila y linia esta escribe el valor correspondiente.
        celda_origen = ws.cell(row=fila_origen, column=col_idx)

        #Copia los formatos de la última linia en las linias importadas
        celda_destino.number_format = celda_origen.number_format
        celda_destino.font = copy(celda_origen.font)
        celda_destino.alignment = copy(celda_origen.alignment)
        celda_destino.border = copy(celda_origen.border)
        celda_destino.fill = copy(celda_origen.fill)
        celda_destino.protection = copy(celda_origen.protection)


def archivar_archivo_importado(archivo: Path, carpeta_archivados: Path) -> None:
    """Mueve el archivo importado a la carpeta de archivados."""
    carpeta_archivados.mkdir(parents=True, exist_ok=True)
    destino = carpeta_archivados / archivo.name
    if archivo.exists() and archivo != destino:
        archivo.replace(destino)


def anadir_movimientos_al_historico(config: dict, filas_nuevas: list) -> int:
    """Añade al final del histórico solo las filas desde la primera UID nueva.
    Devuelve el nº de filas añadidas.

    Compara los UIDs del archivo importado en orden con los UIDs existentes en
    BD_Banco hasta encontrar la primera no coincidente. A partir de esa fila se
    importan todas las siguientes en ambas hojas.
    """
    uids_existentes = leer_uids_existentes(config)

    # Este bloque se conserva comentado como referencia futura para comprobar
    # manualmente valores concretos de UID en BD_Banco A2260 sin afectar la
    # ejecución normal del script.
    # wb_debug = load_workbook(config["ruta_historico"], read_only=True, data_only=True)
    # ws_debug = wb_debug[config["hoja_historico"]]
    # uid_a2260 = ws_debug.cell(row=2260, column=1).value
    # wb_debug.close()
    # print(f"[DEBUG] UID en BD_Banco A2260: {uid_a2260!r}")

    primer_nuevo_idx = None
    for idx, fila in enumerate(filas_nuevas):
        uid = str(fila[0]).strip()
        if uid not in uids_existentes:
            primer_nuevo_idx = idx
            break

    if primer_nuevo_idx is None:
        print("No hay movimientos nuevos que importar.")
        return 0

    filas_a_importar = filas_nuevas[primer_nuevo_idx:]
    total_filas_importado = len(filas_nuevas)
    start = "\033[1;33m"
    end = "\033[0m"
    # Nota: en algunos PowerShell/terminales de Windows no se renderizan
    # las secuencias ANSI; en ese caso el contenido visible será el texto
    # con los marcadores >>> y <<<.
    print(start + "[DEBUG] >>> PRIMERA UID NUEVA: " + end + str(filas_a_importar[0][0]))
    print(start + "[DEBUG] >>> SE IMPORTARÁN {} de {} filas".format(len(filas_a_importar), total_filas_importado) + end + "  del archivo importado.")

    wb = load_workbook(config["ruta_historico"])
    sheet_BDB = wb[config["hoja_historico"]]
    sheet_mov = wb[config["hoja_movimientos_cuenta"]]

    for fila in filas_a_importar:
        escribir_fila_con_formato(sheet_BDB, fila)
        escribir_fila_con_formato(sheet_mov, fila)

    wb.save(config["ruta_historico"])
    ultima_fila_bdb = sheet_BDB.max_row
    ultima_fecha_bdb = sheet_BDB.cell(row=ultima_fila_bdb, column=2).value
    if isinstance(ultima_fecha_bdb, str):
        fecha_formateada = ultima_fecha_bdb
    else:
        fecha_formateada = pd.to_datetime(ultima_fecha_bdb).strftime('%d/%m/%Y')
    start = "\033[1;33m"
    end = "\033[0m"
    print(start +f"[DEBUG] >>> ÚLTIMA FECHA AÑADIDA"+ end +f" EN {config['hoja_historico']}"+ start + f" --> {fecha_formateada}" + end)
    print(f"Se han importado {len(filas_a_importar)} movimientos nuevos.")
    return len(filas_a_importar)


# ====================================================================
# 5. EJECUCIÓN
# ====================================================================

def main():
    archivo = obtener_primer_excel(CONFIG["carpeta_importar"])
    filas_nuevas = leer_movimientos_bancarios(CONFIG)
    anadir_movimientos_al_historico(CONFIG, filas_nuevas)
    archivar_archivo_importado(archivo, CONFIG["carpeta_archivados"])


if __name__ == "__main__":
    main()