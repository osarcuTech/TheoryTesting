# Codigo generado por CLaude a partir de comentarios + ImportarMovimientos.py
import pandas as pd
from openpyxl import load_workbook
from copy import copy
from pathlib import Path

# ====================================================================
# 1. CONFIGURACIÓN
#    (equivalente al "formulario de variables" del script original)
# ====================================================================

CONFIG = {
    # Archivo histórico donde se acumulan todos los movimientos ya importados
    # (equivalente a idSpreadSheet + gidHojaHistorico)
    "ruta_historico": Path("C:/Users/Oscar Ardevol/Documents/MovimientosBancarios/Norgenic/Historico_py.xlsx"),
    "hoja_historico": "BD_Banco",
    "hoja_movimientos_cuenta": "Movimientos_cuenta_0087231",

    # Carpeta donde dejamos el Excel bruto exportado por el banco
    # (equivalente a idCarpetaDrive)
    "carpeta_importar": Path("C:/Users/Oscar Ardevol/Documents/MovimientosBancarios/Norgenic/€/Por archivar"),
    "carpeta_archivados": Path("C:/Users/Oscar Ardevol/Documents/MovimientosBancarios/Norgenic/€/Archivados"),

    # Fila (1-indexada, como en Sheets/Excel) donde empiezan los datos útiles.
    # Varía entre bancos -> equivalente a filaInicioDatosImportados
    "fila_inicio_datos": 4,

    # Nº de columnas de origen que nos interesan: Fecha, FechaValor,
    # Movimiento, MasDatos, Importe, Saldo (A:F)
    "num_columnas_origen": 6,
}


# ====================================================================
# 2. GENERACIÓN DE UID
#    Réplica exacta de la lógica del bloque `const uids = inverseBD.map(...)`
# ====================================================================

def formatear_fecha(valor) -> str:
    """Convierte una fecha a texto DD/MM/YYYY.
    Si está vacía devuelve "" (igual que el `row[0] ? ... : ''` original)."""
    if pd.isna(valor) or valor == "":
        return ""
    if isinstance(valor, str):
        try:
            valor = pd.to_datetime(valor, dayfirst=True)
        except Exception:
            # Si no es una fecha reconocible, lo dejamos tal cual (best-effort,
            # igual que haría Apps Script con new Date() sobre texto raro).
            return valor
    return f"{valor.day:02d}/{valor.month:02d}/{valor.year}"


def formatear_importe(valor) -> str:
    r"""Réplica de: row.toString().replace(/,/g,'').replace(/\./g,',').
    En JavaScript un número entero 1414.0 se convierte en "1414".
    """
    if pd.isna(valor) or valor == "":
        return ""

    if isinstance(valor, (int, float)):
        if float(valor).is_integer():
            return str(int(valor))
        return str(valor).replace(".", ",")

    texto = str(valor).strip()
    if texto == "":
        return ""

    texto_sin_miles = texto.replace(",", "")
    try:
        valor_float = float(texto_sin_miles)
        if valor_float.is_integer():
            return str(int(valor_float))
    except ValueError:
        pass

    return texto.replace(".", ",")


def generar_uid(fila: list) -> str:
    """Genera el UID concatenando los 6 campos con el separador "'_'"."""
    fecha = formatear_fecha(fila[0])
    fecha_valor = formatear_fecha(fila[1])
    movimiento = str(fila[2]).strip() if not pd.isna(fila[2]) else ""
    mas_datos = str(fila[3]).strip() if not pd.isna(fila[3]) else ""
    importe = formatear_importe(fila[4])
    saldo = formatear_importe(fila[5])
    return "'_'".join([fecha, fecha_valor, movimiento, mas_datos, importe, saldo])


# ====================================================================
# 3. LECTURA DEL ARCHIVO BANCARIO BRUTO
#    Equivalente a getMovimientosBancarios()
# ====================================================================

def obtener_primer_excel(carpeta: Path) -> Path:
    """Busca el primer .xlsx o .xls en la carpeta de importación.
    Equivalente a folderMovimientosBancarios.getFilesByType(...).next()"""
    archivos = sorted(
        [p for p in carpeta.iterdir() if p.suffix.lower() in {".xlsx", ".xls"}]
    )
    if not archivos:
        raise FileNotFoundError(f"No se encontró ningún .xlsx/.xls en {carpeta}")
    return archivos[0]


def leer_movimientos_bancarios(config: dict) -> list:
    """Lee el archivo bruto, invierte el orden, filtra filas vacías y genera
    el UID de cada movimiento.
    Devuelve una lista de filas: [UID, Fecha, FechaValor, Movimiento, MasDatos, Importe, Saldo]
    """
    archivo = obtener_primer_excel(config["carpeta_importar"])
    print(f"Leyendo archivo: {archivo.name}")

    # header=None: no confiamos en que la fila de cabecera del banco sea fiable,
    # igual que el original que empieza a leer directamente en filaInicioDatosImportados.
    engine = "openpyxl" if archivo.suffix.lower() == ".xlsx" else "xlrd"
    df = pd.read_excel(
        archivo,
        header=None,
        skiprows=config["fila_inicio_datos"] - 1,
        engine=engine,
    )

    # Nos quedamos solo con las columnas que nos interesan (A:F),
    # por si el banco añade columnas extra a la derecha.
    df = df.iloc[:, :config["num_columnas_origen"]]

    # Filtramos filas totalmente vacías / sin fecha en la columna A
    # (equivalente a: inverseBD.filter(row => row[0] !== ""))
    df = df[df.iloc[:, 0].notna()]

    filas = df.values.tolist()

    # Invertimos el orden (el banco exporta normalmente del más reciente
    # al más antiguo; queremos guardar en orden cronológico ascendente)
    filas.reverse()

    # Generamos el UID y lo anteponemos a cada fila
    filas_con_uid = [[generar_uid(fila)] + fila for fila in filas]

    return filas_con_uid


# ====================================================================
# 3b. ESCENARIO 5: RENOMBRADO DE UID's REPETIDAS + VALIDACIÓN DE SOLAPE
#
#   si( firstDayImportados < LastDayHistorico;
#       ( Renombrar UID's repetidas en Importados as UID_i
#         & ejecutar función python actual );
#       ( print("La fecha de inicio de los importados debe ser anterior a la
#                fecha final del histórico para evitar errores")
#         & break )
#   )
# ====================================================================

def renombrar_uids_repetidos(filas: list) -> list:
    """Si dentro del PROPIO lote importado hay UID's idénticas (ej: dos cargos
    del mismo proveedor, mismo día, mismo importe), les añade un sufijo "_i"
    según el orden de aparición para que no se traten como si fueran la misma
    transacción (Escenario 0/2).
    Ej: 'A', 'A' -> 'A', 'A_2'
    Nota: si más adelante se necesita una columna separada para este sufijo
    (para no romper el SPLIT por "'_'" en otras hojas), esa columna se añade
    manualmente en Excel/Sheets; no es responsabilidad de este script.
    """
    contador = {}
    filas_renombradas = []
    for fila in filas:
        uid_base = fila[0]
        contador[uid_base] = contador.get(uid_base, 0) + 1
        ocurrencia = contador[uid_base]
        uid_final = uid_base if ocurrencia == 1 else f"{uid_base}_{ocurrencia}"
        filas_renombradas.append([uid_final] + fila[1:])
    return filas_renombradas


def obtener_primer_dia_importados(filas: list):
    """Fecha (datetime) del primer movimiento importado (el más antiguo,
    tras haber invertido el orden en leer_movimientos_bancarios)."""
    fecha_texto = filas[0][1]  # columna Fecha, formato DD/MM/YYYY
    return pd.to_datetime(fecha_texto, dayfirst=True)


def obtener_ultimo_dia_historico(config: dict):
    """Fecha (datetime) del último movimiento registrado en BD_Banco."""
    wb = load_workbook(config["ruta_historico"], read_only=True, data_only=True)
    ws = wb[config["hoja_historico"]]
    fecha_valor = ws.cell(row=ws.max_row, column=2).value  # columna B = Fecha
    wb.close()
    if isinstance(fecha_valor, str):
        return pd.to_datetime(fecha_valor, dayfirst=True)
    return pd.to_datetime(fecha_valor)


def validar_solape_suficiente(config: dict, filas_nuevas: list) -> bool:
    """Comprueba que el primer día importado sea ANTERIOR (estrictamente) al
    último día del histórico. Sin ese margen no hay solape que permita
    detectar duplicados de forma fiable, así que se aborta la importación
    en vez de arriesgarse a perder o duplicar movimientos."""
    primer_dia_importados = obtener_primer_dia_importados(filas_nuevas)
    ultimo_dia_historico = obtener_ultimo_dia_historico(config)

    if primer_dia_importados < ultimo_dia_historico:
        return True

    print("La fecha de inicio de los importados debe ser anterior a la fecha "
          "final del histórico para evitar errores.")
    return False


# ====================================================================
# 4. COMPARACIÓN CON EL HISTÓRICO Y CARGA DE MOVIMIENTOS NUEVOS
#    Equivalente a appendBD()
# ====================================================================

def leer_uids_existentes(config: dict) -> set:
    """Lee la columna UID (columna A) del histórico para poder detectar
    duplicados, igual que bdSheetRangeUIDs + historicoUIDs en el original."""
    wb = load_workbook(config["ruta_historico"], read_only=True, data_only=True)
    ws = wb[config["hoja_historico"]]

    uids = set()
    for fila in ws.iter_rows(min_row=2, max_col=1, values_only=True):  # min_row=2 salta cabecera
        if fila[0] is not None:
            uids.add(str(fila[0]).strip())

    wb.close()
    return uids


def encontrar_siguiente_fila(ws) -> int:
    """Devuelve la fila siguiente a la última fila con contenido."""
    for row_idx in range(ws.max_row, 0, -1):
        hay_datos = any(
            ws.cell(row=row_idx, column=col_idx).value is not None
            and str(ws.cell(row=row_idx, column=col_idx).value).strip() != ""
            for col_idx in range(1, ws.max_column + 1)
        )
        if hay_datos:
            return row_idx + 1
    return 1


def escribir_fila_con_formato(ws, valores: list) -> None:
    """Escribe una fila en la siguiente línea libre y copia el formato de la fila anterior."""
    fila_destino = encontrar_siguiente_fila(ws)
    fila_origen = fila_destino - 1 if fila_destino > 1 else 1

    for col_idx, valor in enumerate(valores, start=1):
        celda_destino = ws.cell(row=fila_destino, column=col_idx)
        celda_destino.value = valor

        celda_origen = ws.cell(row=fila_origen, column=col_idx)
        celda_destino.number_format = celda_origen.number_format
        celda_destino.font = copy(celda_origen.font)
        celda_destino.alignment = copy(celda_origen.alignment)
        celda_destino.border = copy(celda_origen.border)
        celda_destino.fill = copy(celda_origen.fill)
        celda_destino.protection = copy(celda_origen.protection)


def archivar_archivo_importado(archivo: Path, carpeta_archivados: Path) -> None:
    """Mueve el archivo importado a la carpeta de archivados."""
    carpeta_archivados.mkdir(parents=True, exist_ok=True)
    destino = carpeta_archivados / archivo.name
    if archivo.exists() and archivo != destino:
        archivo.replace(destino)


def anadir_movimientos_al_historico(config: dict, filas_nuevas: list) -> int:
    """Añade al final del histórico solo las filas desde la primera UID nueva.
    Devuelve el nº de filas añadidas.

    Compara los UIDs del archivo importado en orden con los UIDs existentes en
    BD_Banco hasta encontrar la primera no coincidente. A partir de esa fila se
    importan todas las siguientes en ambas hojas.
    """
    uids_existentes = leer_uids_existentes(config)

    # Este bloque se conserva comentado como referencia futura para comprobar
    # manualmente valores concretos de UID en BD_Banco A2260 sin afectar la
    # ejecución normal del script.
    # wb_debug = load_workbook(config["ruta_historico"], read_only=True, data_only=True)
    # ws_debug = wb_debug[config["hoja_historico"]]
    # uid_a2260 = ws_debug.cell(row=2260, column=1).value
    # wb_debug.close()
    # print(f"[DEBUG] UID en BD_Banco A2260: {uid_a2260!r}")

    primer_nuevo_idx = None
    for idx, fila in enumerate(filas_nuevas):
        uid = str(fila[0]).strip()
        if uid not in uids_existentes:
            primer_nuevo_idx = idx
            break

    if primer_nuevo_idx is None:
        print("No hay movimientos nuevos que importar.")
        return 0

    filas_a_importar = filas_nuevas[primer_nuevo_idx:]
    total_filas_importado = len(filas_nuevas)
    start = "\033[1;33m"
    end = "\033[0m"
    # Nota: en algunos PowerShell/terminales de Windows no se renderizan
    # las secuencias ANSI; en ese caso el contenido visible será el texto
    # con los marcadores >>> y <<<.
    print(start + "[DEBUG] >>> PRIMERA UID NUEVA: " + end + str(filas_a_importar[0][0]))
    print(start + "[DEBUG] >>> SE IMPORTARÁN {} de {} filas".format(len(filas_a_importar), total_filas_importado) + end + "  del archivo importado.")

    wb = load_workbook(config["ruta_historico"])
    sheet_BDB = wb[config["hoja_historico"]]
    sheet_mov = wb[config["hoja_movimientos_cuenta"]]

    for fila in filas_a_importar:
        escribir_fila_con_formato(sheet_BDB, fila)
        escribir_fila_con_formato(sheet_mov, fila)

    wb.save(config["ruta_historico"])
    ultima_fila_bdb = sheet_BDB.max_row
    ultima_fecha_bdb = sheet_BDB.cell(row=ultima_fila_bdb, column=2).value
    if isinstance(ultima_fecha_bdb, str):
        fecha_formateada = ultima_fecha_bdb
    else:
        fecha_formateada = pd.to_datetime(ultima_fecha_bdb).strftime('%d/%m/%Y')
    start = "\033[1;33m"
    end = "\033[0m"
    print(start +f"[DEBUG] >>> ÚLTIMA FECHA AÑADIDA"+ end +f" EN {config['hoja_historico']}"+ start + f" --> {fecha_formateada}" + end)
    print(f"Se han importado {len(filas_a_importar)} movimientos nuevos.")
    return len(filas_a_importar)


# ====================================================================
# 5. EJECUCIÓN
# ====================================================================

def main():
    archivo = obtener_primer_excel(CONFIG["carpeta_importar"])
    filas_nuevas = leer_movimientos_bancarios(CONFIG)

    # Escenario 5: sin solape suficiente no seguimos (evita duplicar o
    # perder movimientos por no poder desambiguar correctamente).
    if not validar_solape_suficiente(CONFIG, filas_nuevas):
        return

    filas_nuevas = renombrar_uids_repetidos(filas_nuevas)

    anadir_movimientos_al_historico(CONFIG, filas_nuevas)
    archivar_archivo_importado(archivo, CONFIG["carpeta_archivados"])


if __name__ == "__main__":
    main()
